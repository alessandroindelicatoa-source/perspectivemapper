"""
PerspectiveMapper - Advanced Discourse Analysis Tool
Supports: TXT, DOCX, PDF, XLSX file formats
Features: Topic modeling, sentiment analysis, bias detection, similarity analysis
"""

import os
import re
import json
import warnings
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from io import BytesIO

import numpy as np
import pandas as pd
import streamlit as st

# Configure page FIRST
st.set_page_config(
    page_title="PerspectiveMapper",
    page_icon="🧭",
    layout="wide",
    initial_sidebar_state="expanded"
)

# NLP / ML
from sklearn.decomposition import PCA
from sklearn.cluster import KMeans
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer
from sklearn.decomposition import LatentDirichletAllocation
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.preprocessing import StandardScaler

# Tokenization & Stopwords
import stopwordsiso
from langdetect import detect, DetectorFactory
DetectorFactory.seed = 0

# WordCloud & Viz
from wordcloud import WordCloud
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.figure_factory as ff
import plotly.graph_objects as go

# File handling
from docx import Document as DocxDocument
import openpyxl

# Sentence embeddings
warnings.filterwarnings("ignore", category=FutureWarning)
from sentence_transformers import SentenceTransformer

# Sentiment analysis
cardiff_error = None
try:
    from transformers import AutoTokenizer, AutoModelForSequenceClassification
    import torch
    CARDIFF_MODEL_NAME = "cardiffnlp/twitter-xlm-roberta-base-sentiment"
    _cardiff_tokenizer = AutoTokenizer.from_pretrained(CARDIFF_MODEL_NAME)
    _cardiff_model = AutoModelForSequenceClassification.from_pretrained(CARDIFF_MODEL_NAME)
    _use_cardiff = True
except Exception as e:
    cardiff_error = str(e)
    _use_cardiff = False
    from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
    _vader_analyzer = SentimentIntensityAnalyzer()

# PDF support
try:
    import pdfplumber
    _pdf_available = True
except ImportError:
    _pdf_available = False


# ========================================
# PASSWORD GATE
# ========================================
def gate():
    """Simple username/password gate using st.secrets['passwords']"""
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if st.session_state.authenticated:
        return True

    st.title("🔐 PerspectiveMapper – Access")
    st.write("Enter your username and password.")
    user = st.text_input("Username", key="auth_user")
    pwd = st.text_input("Password", type="password", key="auth_pwd")

    if st.button("Log in"):
        try:
            if "passwords" in st.secrets and user in st.secrets["passwords"]:
                if str(st.secrets["passwords"][user]) == str(pwd):
                    st.session_state.authenticated = True
                    st.success("Access granted ✅")
                    st.rerun()
                else:
                    st.error("Incorrect password.")
            else:
                st.error("User not found in [passwords].")
        except Exception as e:
            st.error(f"Could not validate access: {e}")
    st.stop()


# ========================================
# UTILITIES
# ========================================
@st.cache_resource(show_spinner=False)
def get_embedder(model_name: str = "paraphrase-multilingual-MiniLM-L12-v2"):
    """Load sentence transformer model with caching"""
    return SentenceTransformer(model_name)


def guess_lang(text: str) -> str:
    """Detect language from text"""
    try:
        return detect(text[:500]) if text else "en"
    except Exception:
        return "en"


def collect_stopwords(selected_langs: List[str], extra_stop: List[str]) -> set:
    """Collect stopwords from multiple languages"""
    sw = set()
    for lang in selected_langs:
        try:
            sw |= set(stopwordsiso.stopwords(lang))
        except Exception:
            pass
    sw |= set([w.strip().lower() for w in extra_stop if w.strip()])
    sw |= set(["http", "https", "www", "com", "url", "link"])
    return sw


def simple_tokenize(text: str) -> List[str]:
    """Tokenize text with URL removal"""
    text = re.sub(r"http\S+|www\.\S+", " ", text, flags=re.I)
    tokens = re.findall(r"[A-Za-zÀ-ÿ0-9_]+", text.lower(), flags=re.U)
    return tokens


def preprocess_docs(docs: List[Dict], stop_set: set) -> List[str]:
    """Preprocess documents by removing stopwords and short tokens"""
    cleaned = []
    for d in docs:
        toks = [t for t in simple_tokenize(d["text"]) if t not in stop_set and len(t) > 2]
        cleaned.append(" ".join(toks))
    return cleaned


def read_txt_file(upload) -> str:
    """Read text file"""
    try:
        return upload.read().decode("utf-8", errors="ignore")
    except Exception as e:
        st.warning(f"Error reading TXT file: {e}")
        return ""


def read_docx_file(upload) -> str:
    """Read DOCX file"""
    try:
        doc = DocxDocument(upload)
        return "\n".join([p.text for p in doc.paragraphs])
    except Exception as e:
        st.warning(f"Error reading DOCX file: {e}")
        return ""


def read_pdf_file(upload) -> str:
    """Read PDF file using pdfplumber"""
    if not _pdf_available:
        st.warning("pdfplumber not installed. Cannot read PDF files.")
        return ""
    try:
        with pdfplumber.open(upload) as pdf:
            text = ""
            for page in pdf.pages:
                extracted = page.extract_text()
                if extracted:
                    text += extracted + "\n"
            return text
    except Exception as e:
        st.warning(f"Error reading PDF file: {e}")
        return ""


def read_excel_file(upload) -> str:
    """Read Excel file and extract text from all cells"""
    try:
        workbook = openpyxl.load_workbook(upload)
        text = ""
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            text += f"\n[Sheet: {sheet}]\n"
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        text += str(cell) + " "
        return text
    except Exception as e:
        st.warning(f"Error reading Excel file: {e}")
        return ""


def read_file(upload) -> str:
    """Read file based on extension"""
    name = upload.name.lower()
    if name.endswith(".txt"):
        return read_txt_file(upload)
    elif name.endswith(".docx"):
        return read_docx_file(upload)
    elif name.endswith(".pdf"):
        return read_pdf_file(upload)
    elif name.endswith((".xlsx", ".xls")):
        return read_excel_file(upload)
    else:
        st.warning(f"Unsupported file type: {name}")
        return ""


def make_wordcloud(text: str, title: str = ""):
    """Generate and display wordcloud"""
    if not text.strip():
        st.caption("Empty text - cannot generate wordcloud")
        return
    try:
        wc = WordCloud(width=1000, height=600, background_color="white").generate(text)
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.imshow(wc, interpolation="bilinear")
        ax.axis("off")
        if title:
            fig.suptitle(title, fontsize=14, fontweight="bold")
        st.pyplot(fig, clear_figure=True)
    except Exception as e:
        st.warning(f"Error generating wordcloud: {e}")


def top_words_per_topic(lda, feature_names, n_top: int = 10) -> pd.DataFrame:
    """Extract top words for each LDA topic"""
    topics = []
    for topic_idx, topic in enumerate(lda.components_):
        top_indices = topic.argsort()[-n_top:][::-1]
        words = [feature_names[i] for i in top_indices]
        scores = [topic[i] for i in top_indices]
        topics.append({
            "topic": topic_idx,
            "top_words": ", ".join(words),
            "avg_score": float(np.mean(scores))
        })
    return pd.DataFrame(topics)


def run_cardiff_sentiment(texts: List[str]) -> Tuple[List[str], List[float]]:
    """Run CardiffNLP sentiment analysis"""
    labels, scores = [], []
    if not _use_cardiff:
        return labels, scores
    id2label = {0: "negative", 1: "neutral", 2: "positive"}
    for t in texts:
        try:
            inputs = _cardiff_tokenizer(t, return_tensors="pt", truncation=True, max_length=256)
            with torch.no_grad():
                outputs = _cardiff_model(**inputs)
                probs = torch.nn.functional.softmax(outputs.logits, dim=-1).cpu().numpy()[0]
            idx = int(np.argmax(probs))
            labels.append(id2label[idx])
            scores.append(float(probs[idx]))
        except Exception:
            labels.append("neutral")
            scores.append(0.0)
    return labels, scores


def run_vader_sentiment(texts: List[str]) -> Tuple[List[str], List[float]]:
    """Run VADER sentiment analysis"""
    from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
    an = SentimentIntensityAnalyzer()
    labels, scores = [], []
    for t in texts:
        res = an.polarity_scores(t)
        comp = res["compound"]
        if comp >= 0.05:
            labels.append("positive")
        elif comp <= -0.05:
            labels.append("negative")
        else:
            labels.append("neutral")
        scores.append(float(comp))
    return labels, scores


def build_bias_table(texts: List[str], bias_dict: Dict[str, List[str]]) -> pd.DataFrame:
    """Analyze bias indicators in texts"""
    rows = []
    lowered = [t.lower() for t in texts]
    for i, t in enumerate(lowered):
        row = {"doc_id": i}
        total = max(len(t.split()), 1)
        for label, kws in bias_dict.items():
            hits = 0
            for kw in kws:
                kw = kw.lower().strip()
                if not kw:
                    continue
                hits += len(re.findall(rf"\b{re.escape(kw)}\b", t))
            row[label] = hits / total
        rows.append(row)
    return pd.DataFrame(rows)


def calculate_document_statistics(docs: List[Dict], cleaned: List[str]) -> pd.DataFrame:
    """Calculate detailed statistics for each document"""
    stats = []
    for i, (doc, clean_text) in enumerate(zip(docs, cleaned)):
        raw_text = doc["text"]
        stats.append({
            "document": doc["filename"],
            "language": doc["lang"],
            "raw_chars": len(raw_text),
            "raw_words": len(raw_text.split()),
            "cleaned_words": len(clean_text.split()),
            "unique_words": len(set(clean_text.split())),
            "avg_word_length": np.mean([len(w) for w in clean_text.split()]) if clean_text.split() else 0,
            "sentences": len(re.split(r'[.!?]+', raw_text)),
        })
    return pd.DataFrame(stats)


def calculate_tfidf_analysis(cleaned: List[str], feature_names: List[str]) -> pd.DataFrame:
    """Calculate TF-IDF scores for top terms"""
    try:
        vectorizer = TfidfVectorizer(max_features=50)
        tfidf_matrix = vectorizer.fit_transform(cleaned)
        feature_names_tfidf = vectorizer.get_feature_names_out()
        
        # Get top terms by average TF-IDF score
        avg_tfidf = np.asarray(tfidf_matrix.mean(axis=0)).flatten()
        top_indices = avg_tfidf.argsort()[-20:][::-1]
        
        results = []
        for idx in top_indices:
            results.append({
                "term": feature_names_tfidf[idx],
                "avg_tfidf": float(avg_tfidf[idx])
            })
        return pd.DataFrame(results)
    except Exception as e:
        st.warning(f"Error calculating TF-IDF: {e}")
        return pd.DataFrame()


# ========================================
# MAIN APP
# ========================================
gate()  # Require password

# Header
c1, c2 = st.columns([1, 4])
with c1:
    logo_candidates = [
        Path(__file__).parent / "assets" / "logo.png",
        Path.cwd() / "assets" / "logo.png",
        "assets/logo.png",
    ]
    logo_path = None
    for p in logo_candidates:
        try:
            if isinstance(p, str):
                if os.path.exists(p):
                    logo_path = p
                    break
            else:
                if p.exists():
                    logo_path = str(p)
                    break
        except Exception:
            pass
    if logo_path:
        st.image(logo_path, use_container_width=True)

with c2:
    st.title("PerspectiveMapper")
    st.write("Advanced discourse analysis: topics, sentiment, similarity, and bias detection.")

# Sidebar Configuration
st.sidebar.header("⚙️ Settings")

# File upload with enhanced format support
file_types = ["txt", "docx", "pdf", "xlsx", "xls"]
uploads = st.sidebar.file_uploader(
    "Upload files (.txt, .docx, .pdf, .xlsx)",
    type=file_types,
    accept_multiple_files=True
)

# Analysis settings
st.sidebar.subheader("Analysis Options")
lang_codes = st.sidebar.multiselect(
    "Stopword languages",
    ["en", "es", "it", "fr", "de", "pt", "ca", "eu", "gl"],
    default=["en", "es", "it"]
)
extra_sw = st.sidebar.text_area(
    "Extra stopwords (comma-separated)",
    value="and, the, of, to, a, in, is, it"
)

# Visualization options
st.sidebar.subheader("Visualization Options")
show_wc = st.sidebar.checkbox("Show WordCloud per document", value=True)
show_stats = st.sidebar.checkbox("Show document statistics", value=True)
show_tfidf = st.sidebar.checkbox("Show TF-IDF analysis", value=True)

# Analysis parameters
st.sidebar.subheader("Model Parameters")
n_topics = st.sidebar.slider("LDA topics", 2, 12, 5)
max_features = st.sidebar.slider("Max vocabulary size", 1000, 10000, 3000, step=500)
n_clusters = st.sidebar.slider("KMeans clusters", 2, 12, 4)

# Sentiment analysis
st.sidebar.subheader("Sentiment Analysis")
use_cardiff = st.sidebar.checkbox("Use CardiffNLP sentiment", value=True)

# Bias detection
st.sidebar.subheader("Bias Detection")
bias_json = st.sidebar.text_area(
    "Bias dictionary JSON",
    value=json.dumps({
        "gender": ["woman", "man", "trans", "equality", "female", "male"],
        "migration": ["immigrant", "migrant", "refugee", "border", "asylum"],
        "religion": ["church", "islam", "catholic", "jewish", "christian"],
        "politics": ["left", "right", "liberal", "conservative", "democrat", "republican"]
    }, indent=2)
)

# Export options
st.sidebar.subheader("Export Options")
want_csv = st.sidebar.checkbox("Enable CSV download", value=True)
want_json = st.sidebar.checkbox("Enable JSON download", value=True)

# Main Content
if not uploads:
    st.info("⬅️ Upload one or more files to begin the analysis.")
    st.stop()

# Load documents
docs = []
for up in uploads:
    text = read_file(up)
    if text.strip():
        lang = guess_lang(text)
        docs.append({"filename": up.name, "text": text, "lang": lang})
    else:
        st.warning(f"⚠️ {up.name} is empty or could not be read.")

if not docs:
    st.error("No valid documents loaded. Please check your files.")
    st.stop()

st.success(f"✅ Loaded {len(docs)} document(s).")

# Preprocessing
stop_set = collect_stopwords(lang_codes, [w.strip() for w in extra_sw.split(",")])
cleaned = preprocess_docs(docs, stop_set)

# ========================================
# DOCUMENT STATISTICS
# ========================================
if show_stats:
    st.subheader("📊 Document Statistics")
    stats_df = calculate_document_statistics(docs, cleaned)
    st.dataframe(stats_df, use_container_width=True)

# ========================================
# WORDCLOUDS
# ========================================
if show_wc:
    st.subheader("☁️ WordCloud per document")
    for d, text in zip(docs, cleaned):
        st.markdown(f"**{d['filename']}**")
        make_wordcloud(text, title=d['filename'])

# ========================================
# TF-IDF ANALYSIS
# ========================================
if show_tfidf:
    st.subheader("🔤 TF-IDF Analysis (Top Terms)")
    tfidf_df = calculate_tfidf_analysis(cleaned, [])
    if not tfidf_df.empty:
        col1, col2 = st.columns([2, 1])
        with col1:
            fig = px.bar(
                tfidf_df,
                x="avg_tfidf",
                y="term",
                orientation="h",
                title="Top 20 Terms by TF-IDF Score",
                labels={"avg_tfidf": "Average TF-IDF Score", "term": "Term"}
            )
            st.plotly_chart(fig, use_container_width=True)
        with col2:
            st.dataframe(tfidf_df, use_container_width=True)

# ========================================
# TOPIC MODELING (LDA)
# ========================================
st.subheader("🧵 Topic Modeling (LDA)")
try:
    vectorizer = CountVectorizer(max_features=max_features)
    X = vectorizer.fit_transform(cleaned)
    lda = LatentDirichletAllocation(n_components=n_topics, random_state=42, max_iter=20)
    W = lda.fit_transform(X)
    topics_df = top_words_per_topic(lda, vectorizer.get_feature_names_out(), n_top=12)
    st.dataframe(topics_df, use_container_width=True)
    dominant_topic = np.argmax(W, axis=1).tolist()
except Exception as e:
    st.error(f"Error in LDA: {e}")
    st.stop()

# ========================================
# CLUSTERING (PCA + KMeans)
# ========================================
st.subheader("🧭 Clustering (SBERT + PCA + KMeans)")
try:
    embedder = get_embedder()
    embeddings = embedder.encode([d["text"] for d in docs], show_progress_bar=False)
    n_docs = len(docs)
    
    if n_docs < 2:
        coords = np.zeros((n_docs, 2))
    else:
        n_comp = min(2, n_docs, embeddings.shape[1])
        pca = PCA(n_components=n_comp, random_state=42)
        coords_pca = pca.fit_transform(embeddings)
        coords = coords_pca if coords_pca.shape[1] == 2 else np.hstack([coords_pca, np.zeros((n_docs, 1))])
    
    df_plot = pd.DataFrame({
        "x": coords[:, 0],
        "y": coords[:, 1],
        "file": [d["filename"] for d in docs],
        "topic": [f"T{t}" for t in dominant_topic]
    })
    
    k_for_fit = min(max(1, n_clusters), n_docs)
    clusters = np.zeros(n_docs, dtype=int) if k_for_fit < 2 else KMeans(
        n_clusters=k_for_fit, random_state=42, n_init="auto"
    ).fit_predict(coords)
    df_plot["cluster"] = clusters
    
    fig = px.scatter(
        df_plot,
        x="x",
        y="y",
        text="file",
        color="cluster",
        title="Document Clustering (PCA Projection)",
        labels={"x": "PC1", "y": "PC2"}
    )
    fig.update_traces(textposition="top center")
    st.plotly_chart(fig, use_container_width=True)
except Exception as e:
    st.error(f"Error in clustering: {e}")
    clusters = np.zeros(n_docs, dtype=int)

# ========================================
# HIERARCHICAL CLUSTERING DENDROGRAM
# ========================================
st.subheader("🌳 Hierarchical Clustering Dendrogram")
try:
    import scipy.cluster.hierarchy as sch
    from scipy.spatial.distance import pdist
    
    dist_matrix = pdist(embeddings, metric="cosine")
    linkage = sch.linkage(dist_matrix, method="ward")
    
    fig, ax = plt.subplots(figsize=(12, 6))
    sch.dendrogram(
        linkage,
        labels=[d["filename"] for d in docs],
        orientation="top",
        leaf_rotation=45,
        leaf_font_size=10,
        ax=ax,
    )
    ax.set_title("Hierarchical Clustering of Documents")
    ax.set_ylabel("Distance")
    st.pyplot(fig)
except Exception as e:
    st.warning(f"Error generating dendrogram: {e}")

# ========================================
# SIMILARITY MATRIX
# ========================================
st.subheader("🔗 Document Similarity Matrix")
try:
    sim = cosine_similarity(embeddings)
    heatmap = ff.create_annotated_heatmap(
        z=sim,
        x=[d["filename"] for d in docs],
        y=[d["filename"] for d in docs],
        showscale=True,
        colorscale="Viridis"
    )
    heatmap.update_layout(title="Cosine Similarity Between Documents")
    st.plotly_chart(heatmap, use_container_width=True)
except Exception as e:
    st.error(f"Error generating similarity matrix: {e}")

# ========================================
# SENTIMENT ANALYSIS
# ========================================
st.subheader("💬 Sentiment Analysis")
try:
    texts = [d["text"] for d in docs]
    if use_cardiff and _use_cardiff:
        labels, scores = run_cardiff_sentiment(texts)
    else:
        labels, scores = run_vader_sentiment(texts)
    
    sentiment_df = pd.DataFrame({
        "document": [d["filename"] for d in docs],
        "sentiment": labels,
        "score": scores
    })
    
    col1, col2 = st.columns([2, 1])
    with col1:
        fig = px.bar(
            sentiment_df,
            x="document",
            y="score",
            color="sentiment",
            title="Sentiment Scores by Document",
            labels={"score": "Sentiment Score", "document": "Document"}
        )
        st.plotly_chart(fig, use_container_width=True)
    with col2:
        st.dataframe(sentiment_df, use_container_width=True)
except Exception as e:
    st.error(f"Error in sentiment analysis: {e}")
    sentiment_df = pd.DataFrame()

# ========================================
# BIAS ANALYSIS
# ========================================
st.subheader("🧷 Bias Indicator Analysis")
try:
    bias_dict = json.loads(bias_json)
except Exception as e:
    st.error(f"Invalid JSON for bias keywords: {e}")
    bias_dict = {}

bias_df = build_bias_table(texts, bias_dict) if bias_dict else pd.DataFrame()

if not bias_df.empty:
    st.dataframe(bias_df, use_container_width=True)
    
    # Visualization
    fig = px.bar(
        bias_df.melt(id_vars=["doc_id"], var_name="bias_category", value_name="score"),
        x="doc_id",
        y="score",
        color="bias_category",
        title="Bias Indicators by Document",
        labels={"doc_id": "Document", "score": "Bias Score"}
    )
    st.plotly_chart(fig, use_container_width=True)

# ========================================
# COMPREHENSIVE RESULTS TABLE
# ========================================
st.subheader("📈 Comprehensive Results")
results = pd.DataFrame({
    "file": [d["filename"] for d in docs],
    "language": [d["lang"] for d in docs],
    "tokens_len": [len(c.split()) for c in cleaned],
    "lda_dominant_topic": dominant_topic,
    "cluster": clusters,
    "pca_x": coords[:, 0],
    "pca_y": coords[:, 1]
})

if not sentiment_df.empty:
    sentiment_simple = sentiment_df[["document", "sentiment", "score"]].copy()
    sentiment_simple.columns = ["file", "sentiment", "sentiment_score"]
    results = results.merge(sentiment_simple, on="file", how="left")

if not bias_df.empty:
    results = results.merge(bias_df, left_index=True, right_on="doc_id", how="left").drop(columns=["doc_id"])

st.dataframe(results, use_container_width=True)

# ========================================
# EXPORT OPTIONS
# ========================================
st.subheader("⬇️ Download Results")

col1, col2, col3 = st.columns(3)

if want_csv:
    with col1:
        csv_data = results.to_csv(index=False).encode("utf-8")
        st.download_button(
            label="📥 Download CSV",
            data=csv_data,
            file_name="perspectivemapper_results.csv",
            mime="text/csv"
        )

if want_json:
    with col2:
        json_data = json.dumps({
            "summary": {
                "total_documents": len(docs),
                "total_topics": n_topics,
                "total_clusters": k_for_fit
            },
            "documents": results.to_dict(orient="records"),
            "topics": topics_df.to_dict(orient="records") if 'topics_df' in locals() else []
        }, indent=2).encode("utf-8")
        st.download_button(
            label="📥 Download JSON",
            data=json_data,
            file_name="perspectivemapper_results.json",
            mime="application/json"
        )

with col3:
    if st.button("🔄 Reset Analysis"):
        st.session_state.clear()
        st.rerun()

st.markdown("---")
st.caption("PerspectiveMapper v2.0 - Advanced Discourse Analysis Tool")
